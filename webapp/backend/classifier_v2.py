#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
classifier_v2.py
Chromium/web implementation of Artikelplacering with broad parity against classifier.py.
"""

from __future__ import annotations

import asyncio
import csv
import json
import mimetypes
import os
import queue
import random
import re
import shutil
import tempfile
import threading
import time
import uuid
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

import openpyxl
import requests
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, Response, StreamingResponse
from pydantic import BaseModel


router = APIRouter(prefix="/api/classifier/v2", tags=["classifier-v2"])

# Keep these constants aligned with classifier.py where possible.
DEFAULT_AI_URL = "http://localhost:1234/v1"
MAX_EXAMPLES_PER_CAT = 10
EXT_IMAGES_PER_CAT = 10
EXT_BATCH_SIZE = 1
DEFAULT_SYFTE = (
    "Kategorisera artiklar för att stödja pallbyggnation. "
    "Var konsekvent och välj den mest specifika passande kategorin."
)
AI_JOB_MIN_PER_CAT = 0
AI_PARALLEL_WORKERS = 3
CLASSIFIER_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".tiff"}

CLASSIFIER_APP_DIR = Path(
    os.environ.get("CLASSIFIER_APP_DIR", r"C:\artikelplacering\Artikelplacering").strip()
)
DEFAULT_IMAGE_DIR = Path(
    os.environ.get("CLASSIFIER_DEFAULT_IMAGE_DIR", str(CLASSIFIER_APP_DIR / "bilder")).strip()
)
DEFAULT_OUTPUT_DIR = Path(
    os.environ.get("CLASSIFIER_DEFAULT_OUTPUT_DIR", str(CLASSIFIER_APP_DIR)).strip()
)
DEFAULT_DATA_DIR = CLASSIFIER_APP_DIR / "data"
CLASSIFIER_WEB_V2_DATA_DIR = Path(
    os.environ.get(
        "CLASSIFIER_WEB_V2_DATA_DIR",
        str(Path(tempfile.gettempdir()) / "allok_classifier_v2_data"),
    )
)
CLASSIFIER_DATA_FILE_KEYS = {"item", "item_alias", "item_attribute", "main_category"}


def _safe_name(name: str) -> str:
    return re.sub(r"[^\w.\-]+", "_", (name or "").strip())


def _safe_folder_fragment(name: str) -> str:
    text = str(name or "").strip()
    text = re.sub(r'[\\/:*?"<>|]+', "_", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text or "unnamed"


def _ts() -> str:
    return time.strftime("%H:%M:%S")


def _now_token() -> str:
    return time.strftime("%Y%m%d%H%M%S")


def _unique_target_path(dst_dir: Path, filename: str) -> Path:
    target = dst_dir / filename
    if not target.exists():
        return target
    stem = target.stem
    suffix = target.suffix
    for i in range(2, 10000):
        candidate = dst_dir / f"{stem}_{i}{suffix}"
        if not candidate.exists():
            return candidate
    return dst_dir / f"{stem}_{int(time.time())}{suffix}"


def _parse_content_disposition_filename(header_value: str) -> str:
    if not header_value:
        return ""
    m = re.search(r"filename\*=UTF-8''([^;]+)", header_value, flags=re.IGNORECASE)
    if m:
        return requests.utils.unquote(m.group(1).strip().strip('"'))
    m = re.search(r'filename="([^"]+)"', header_value, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r"filename=([^;]+)", header_value, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip().strip('"')
    return ""


def _download_url_bytes(url: str, timeout_sec: int = 45) -> Tuple[bytes, str]:
    resp = requests.get(url, timeout=timeout_sec)
    if not resp.ok or not resp.content:
        raise RuntimeError(f"Kunde inte ladda bild-URL ({resp.status_code}): {url}")
    content_type = (resp.headers.get("content-type") or "").split(";")[0].strip()
    return resp.content, (content_type or "application/octet-stream")


def _filename_from_url_row(row: Dict[str, str], index: int, content_type: str) -> str:
    article = _safe_folder_fragment(str(row.get("article_number", "") or "").strip())
    if not article:
        article = f"image_{index + 1}"
    url_path = urlparse(str(row.get("url", "") or "")).path
    ext = Path(url_path).suffix.lower()
    if ext not in CLASSIFIER_IMAGE_EXTENSIONS:
        guessed = mimetypes.guess_extension(content_type or "")
        ext = (guessed or ".jpg").lower()
    if not ext.startswith("."):
        ext = f".{ext}"
    return f"{article}{ext}"


# ---------------------------------------------------------------------------
# Data manager (ported from classifier.py for parity)
# ---------------------------------------------------------------------------


class DataManager:
    def __init__(self, data_dir: Optional[Path] = None):
        self.data_dir = Path(data_dir) if data_dir else DEFAULT_DATA_DIR
        self.builtin_attributes: List[Dict] = []
        self.store_quantity_data: Dict[Tuple[str, str], str] = {}
        self.item_data: Dict[str, Dict] = {}
        self.alias_data: Dict[str, Dict] = {}
        self.category_map: Dict[str, str] = {}
        self._load_all()

    def _load_all(self):
        if not self.data_dir.exists():
            return
        for f in sorted(self.data_dir.iterdir()):
            name = f.name.lower()
            if not name.endswith(".csv"):
                continue
            if name.startswith("item_attribute"):
                self._load_attributes(f)
            elif name.startswith("item_alias"):
                self._load_alias(f)
            elif name.startswith("item") and not name.startswith("item_"):
                self._load_items(f)
            elif name.startswith("main_category"):
                self._load_main_category(f)

    def _read_tsv(self, path: Path) -> List[Dict]:
        try:
            with open(path, newline="", encoding="utf-8-sig") as fh:
                sample = fh.read(4096)
                fh.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                return list(csv.DictReader(fh, dialect=dialect))
        except Exception:
            return []

    def _load_attributes(self, path: Path):
        self.builtin_attributes = []
        self.store_quantity_data = {}
        art_data: Dict[Tuple[str, str], Dict] = {}
        for row in self._read_tsv(path):
            art = row.get("Artikel", "").strip()
            bolag = row.get("Bolag", "").strip()
            namn = row.get("Namn", "").strip()
            val = row.get("Värde", "").strip()
            if not art:
                continue
            key = (art, bolag)
            if key not in art_data:
                art_data[key] = {"bolag": bolag}
            if namn == "IMG" and val.lower().startswith("http"):
                art_data[key]["url"] = val
            elif namn == "StoreQuantity":
                art_data[key]["store_quantity"] = val
        for (art, bolag), data in art_data.items():
            if "url" in data:
                self.builtin_attributes.append(
                    {"article_number": art, "url": data["url"], "bolag": bolag}
                )
            if "store_quantity" in data:
                self.store_quantity_data[(art, bolag)] = data["store_quantity"]

    def _load_alias(self, path: Path):
        self.alias_data = {}
        for row in self._read_tsv(path):
            art = row.get("Artikel", "").strip()
            if not art or art in self.alias_data:
                continue
            self.alias_data[art] = {
                "ean": row.get("Alias", "").strip(),
                "enhet": row.get("Enhet", "").strip(),
                "faktor": row.get("Faktor", "").strip(),
                "langd": row.get("Längd", "").strip(),
                "bredd": row.get("Bredd", "").strip(),
                "hojd": row.get("Höjd", "").strip(),
                "bolag": row.get("Bolag", "").strip(),
            }

    def _load_items(self, path: Path):
        self.item_data = {}
        for row in self._read_tsv(path):
            art = row.get("Artikel", "").strip()
            if not art:
                continue
            self.item_data[art] = {
                "beskrivning": row.get("Beskrivning", "").strip(),
                "un_nummer": row.get("UN nummer", "").strip(),
                "vikt_brutto": row.get("Vikt brutto", "").strip(),
                "vikt_netto": row.get("Vikt netto", "").strip(),
                "volym": row.get("Volym", "").strip(),
                "kategori": row.get("Kategori", "").strip(),
                "robot": row.get("Robot", "").strip(),
                "bolag": row.get("Bolag", "").strip(),
            }

    def _load_main_category(self, path: Path):
        self.category_map = {}
        for row in self._read_tsv(path):
            kat = row.get("Kategori", "").strip()
            hkat = row.get("Huvudkategori", "").strip()
            if kat and hkat:
                self.category_map[kat] = hkat

    def get_meta(self, article_str: str, bolag: str = "") -> Optional[Dict]:
        art = article_str.strip()
        result: Dict = {}
        if art in self.item_data:
            result.update(self.item_data[art])
        if art in self.alias_data:
            result.update(self.alias_data[art])
        cat_code = result.get("kategori", "")
        if cat_code and cat_code in self.category_map:
            result["huvudkategori"] = self.category_map[cat_code]
        sq = self.store_quantity_data.get((art, bolag))
        if sq is None:
            sq = next((v for (a, _), v in self.store_quantity_data.items() if a == art), None)
        if sq is not None:
            result["store_quantity"] = sq
        return result or None


# ---------------------------------------------------------------------------
# Session state
# ---------------------------------------------------------------------------


@dataclass
class V2Category:
    name: str
    description: str = ""

    def to_dict(self) -> Dict[str, str]:
        return {"name": self.name, "description": self.description}


@dataclass
class V2Session:
    session_id: str
    test_name: str
    syfte: str
    categories: List[V2Category]
    csv_data: List[Dict[str, Any]]
    output_dir: str
    image_dir: str = ""
    image_source: str = "url"  # url|image_dir|mixed
    current_index: int = 0
    skipped: List[str] = field(default_factory=list)
    results: List[Dict[str, Any]] = field(default_factory=list)
    categorized: List[Dict[str, Any]] = field(default_factory=list)
    cat_knowledge: Dict[str, str] = field(default_factory=dict)
    cat_example_articles: Dict[str, List[str]] = field(default_factory=dict)
    stage: str = "classify"
    created_at: float = field(default_factory=time.time)
    updated_at: float = field(default_factory=time.time)
    ai_running: bool = False
    ai_paused: bool = False
    ai_step: str = "idle"  # idle|step1|step1_done|step2|done|stopped|error
    ai_error: str = ""
    ai_progress_done: int = 0
    ai_progress_total: int = 0
    ai_settings: Dict[str, Any] = field(default_factory=dict)
    log_lines: List[str] = field(default_factory=list)
    event_queue: queue.Queue = field(default_factory=queue.Queue)
    temp_dir: str = field(default_factory=lambda: tempfile.mkdtemp(prefix="classifier_v2_"))
    lock: threading.RLock = field(default_factory=threading.RLock)
    ai_stop_event: threading.Event = field(default_factory=threading.Event)
    ai_pause_event: threading.Event = field(default_factory=threading.Event)

    def category_names(self) -> List[str]:
        return [c.name for c in self.categories]


_sessions_lock = threading.RLock()
_sessions: Dict[str, V2Session] = {}
_data_files_lock = threading.RLock()
_data_files: Dict[str, str] = {}
_active_workers_lock = threading.RLock()
_active_workers: Dict[str, "V2AIWorker"] = {}


def _log(session: V2Session, message: str):
    line = f"[{_ts()}] {message}"
    with session.lock:
        session.log_lines.append(line)
        if len(session.log_lines) > 2000:
            session.log_lines = session.log_lines[-1500:]
        session.updated_at = time.time()
    try:
        session.event_queue.put_nowait(line)
    except Exception:
        pass


def _data_files_payload() -> Dict[str, Dict[str, Any]]:
    payload: Dict[str, Dict[str, Any]] = {}
    with _data_files_lock:
        for key in sorted(CLASSIFIER_DATA_FILE_KEYS):
            p = _data_files.get(key, "")
            exists = bool(p and os.path.exists(p))
            payload[key] = {
                "exists": exists,
                "path": p if exists else "",
                "filename": os.path.basename(p) if exists else "",
            }
    return payload


def _resolve_data_dir() -> Path:
    CLASSIFIER_WEB_V2_DATA_DIR.mkdir(parents=True, exist_ok=True)
    return CLASSIFIER_WEB_V2_DATA_DIR


def _current_data_manager() -> DataManager:
    # Build a DataManager from v2 uploaded files if present, else fallback to app data.
    base = _resolve_data_dir()
    with _data_files_lock:
        file_map = dict(_data_files)
    # Mirror file naming expected by DataManager _load_all prefixes.
    for key, src in file_map.items():
        if not src or not os.path.exists(src):
            continue
        ext = Path(src).suffix or ".csv"
        if key == "item":
            target = base / f"item{ext}"
        else:
            target = base / f"{key}{ext}"
        try:
            shutil.copy2(src, target)
        except Exception:
            pass
    if any((base / f).exists() for f in ("item.csv", "item_attribute.csv", "item_alias.csv", "main_category.csv")):
        return DataManager(base)
    return DataManager(DEFAULT_DATA_DIR)


def _normalize_categories(raw_categories: List[Any]) -> List[V2Category]:
    out: List[V2Category] = []
    seen = set()
    for entry in raw_categories or []:
        if isinstance(entry, dict):
            name = str(entry.get("name", "")).strip()
            desc = str(entry.get("description", "")).strip()
        else:
            name = str(entry or "").strip()
            desc = ""
        if not name:
            continue
        low = name.lower()
        if low in seen:
            continue
        seen.add(low)
        out.append(V2Category(name=name, description=desc))
    if "övrigt" not in {c.name.lower() for c in out}:
        out.append(V2Category(name="Övrigt", description="Fallbackkategori"))
    return out


def _extract_rows_from_item_attribute(path: str) -> List[Dict[str, str]]:
    try:
        with open(path, encoding="utf-8-sig", newline="") as fh:
            sample = fh.read(4096)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(fh, dialect=dialect)
            rows = list(reader)
    except Exception:
        return []

    out: List[Dict[str, str]] = []
    seen = set()
    for row in rows:
        art = str(row.get("Artikel", "") or row.get("Artikelnummer", "") or "").strip()
        name = str(row.get("Namn", "") or row.get("Name", "") or "").strip().lower()
        val = str(row.get("Värde", "") or row.get("Varde", "") or row.get("Value", "") or "").strip()
        bolag = str(row.get("Bolag", "") or "").strip()
        if not art or not val.lower().startswith("http"):
            continue
        if name and name not in {"img", "image", "bild", "url"}:
            continue
        key = (art, bolag, val)
        if key in seen:
            continue
        seen.add(key)
        out.append({"article_number": art, "url": val, "bolag": bolag, "img_path": ""})
    return out


def _rows_from_image_dir(image_dir: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    if not image_dir.exists() or not image_dir.is_dir():
        return rows
    files = [p for p in sorted(image_dir.rglob("*")) if p.is_file() and p.suffix.lower() in CLASSIFIER_IMAGE_EXTENSIONS]
    for p in files:
        rows.append(
            {
                "article_number": p.stem,
                "url": "",
                "bolag": "",
                "img_path": str(p.resolve()),
            }
        )
    return rows


def _result_map(session: V2Session) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for row in session.results:
        out[str(row.get("article_number", "")).strip()] = row
    return out


def _unclassified_rows(session: V2Session) -> List[Dict[str, Any]]:
    classified = set(_result_map(session).keys())
    return [r for r in session.csv_data if str(r.get("article_number", "")).strip() not in classified]


def _advance_to_next_unclassified(session: V2Session):
    classified = set(_result_map(session).keys())
    while session.current_index < len(session.csv_data):
        art = str(session.csv_data[session.current_index].get("article_number", "")).strip()
        if art and art not in classified:
            break
        session.current_index += 1


def _current_row(session: V2Session) -> Optional[Dict[str, Any]]:
    if session.current_index < 0 or session.current_index >= len(session.csv_data):
        return None
    return session.csv_data[session.current_index]


def _find_session_or_404(session_id: str) -> V2Session:
    with _sessions_lock:
        session = _sessions.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Klassificeringssession saknas.")
    return session


def _build_kanban(session: V2Session) -> Dict[str, Any]:
    columns: Dict[str, List[Dict[str, Any]]] = {name: [] for name in session.category_names()}
    for res in session.results:
        cat = str(res.get("category", "")).strip()
        if cat not in columns:
            columns.setdefault(cat, [])
        columns[cat].append(
            {
                "article_number": str(res.get("article_number", "")),
                "category": cat,
                "reason": str(res.get("reason", "")),
                "url": str(res.get("url", "")),
                "img_path": str(res.get("image_path", "")),
                "bolag": str(res.get("bolag", "")),
            }
        )
    for cards in columns.values():
        cards.sort(key=lambda c: c.get("article_number", ""))
    return {
        "columns": [
            {
                "category": name,
                "knowledge_ready": bool(session.cat_knowledge.get(name)),
                "cards": columns.get(name, []),
            }
            for name in session.category_names()
        ]
    }


def _session_state(session: V2Session) -> Dict[str, Any]:
    with session.lock:
        _advance_to_next_unclassified(session)
        current = _current_row(session)
        total = len(session.csv_data)
        done = len(_result_map(session)) >= total and total > 0
        if done:
            session.stage = "done"
        counts: Dict[str, int] = {name: 0 for name in session.category_names()}
        for res in session.results:
            cat = str(res.get("category", "")).strip()
            counts[cat] = int(counts.get(cat, 0)) + 1
        out = {
            "ok": True,
            "session_id": session.session_id,
            "stage": session.stage,
            "test_name": session.test_name,
            "syfte": session.syfte,
            "categories": [c.to_dict() for c in session.categories],
            "image_dir": session.image_dir,
            "output_dir": session.output_dir,
            "image_source": session.image_source,
            "total": total,
            "classified": len(session.results),
            "unclassified": max(0, total - len(session.results)),
            "skipped_count": len(session.skipped),
            "index": session.current_index,
            "done": done,
            "counts": counts,
            "current": current or {},
            "cat_knowledge": dict(session.cat_knowledge),
            "cat_example_articles": dict(session.cat_example_articles),
            "ai": {
                "running": session.ai_running,
                "paused": session.ai_paused,
                "step": session.ai_step,
                "error": session.ai_error,
                "progress_done": session.ai_progress_done,
                "progress_total": session.ai_progress_total,
            },
            "kanban": _build_kanban(session),
            "logs_tail": session.log_lines[-200:],
            "created_at": session.created_at,
            "updated_at": session.updated_at,
        }
    return out


def _copy_classified_file(session: V2Session, row: Dict[str, Any], category: str) -> str:
    output_dir = Path(session.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    dst_dir = output_dir / f"{_safe_folder_fragment(session.test_name)}.{_safe_folder_fragment(category)}"
    dst_dir.mkdir(parents=True, exist_ok=True)
    img_path = str(row.get("img_path", "") or "").strip()

    if img_path and Path(img_path).exists():
        src = Path(img_path)
        dst = _unique_target_path(dst_dir, src.name)
        shutil.copy2(src, dst)
        return str(dst)

    url = str(row.get("url", "") or "").strip()
    if not url:
        return ""
    content, content_type = _download_url_bytes(url, timeout_sec=45)
    fname = _filename_from_url_row(row, 0, content_type)
    dst = _unique_target_path(dst_dir, fname)
    with open(dst, "wb") as fh:
        fh.write(content)
    return str(dst)


def _set_result(session: V2Session, row: Dict[str, Any], category: str, reason: str = "", source: str = "manual"):
    art = str(row.get("article_number", "")).strip()
    if not art:
        return
    if category not in session.category_names():
        category = "Övrigt"

    saved_path = ""
    try:
        saved_path = _copy_classified_file(session, row, category)
    except Exception as e:
        _log(session, f"Varning: kunde inte kopiera/spara bild för {art}: {e}")

    res_map = _result_map(session)
    payload = {
        "article_number": art,
        "category": category,
        "reason": reason or "",
        "url": str(row.get("url", "") or ""),
        "bolag": str(row.get("bolag", "") or ""),
        "image_path": saved_path or str(row.get("img_path", "") or ""),
        "source": source,
    }
    res_map[art] = payload
    session.results = list(res_map.values())

    cat_map: Dict[str, Dict[str, Any]] = {
        str(x.get("article_number", "")).strip(): x for x in session.categorized
    }
    cat_map[art] = {
        "article_number": art,
        "category": category,
        "image_path": payload["image_path"],
    }
    session.categorized = list(cat_map.values())
    session.updated_at = time.time()


# ---------------------------------------------------------------------------
# AI worker
# ---------------------------------------------------------------------------


def _call_chat_completion(
    api_url: str,
    model: str,
    messages: List[Dict[str, str]],
    api_key: str = "",
    timeout_sec: int = 90,
    retries: int = 3,
) -> str:
    base = (api_url or DEFAULT_AI_URL).rstrip("/")
    url = f"{base}/chat/completions"
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    payload = {
        "model": model,
        "messages": messages,
        "temperature": 0.2,
    }
    last_err = "okänt fel"
    for _ in range(max(1, retries)):
        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=timeout_sec)
            if not resp.ok:
                try:
                    detail = resp.json()
                except Exception:
                    detail = resp.text[:400]
                last_err = f"{resp.status_code}: {detail}"
                time.sleep(1.0)
                continue
            data = resp.json()
            choices = data.get("choices") or []
            if not choices:
                last_err = "Saknar choices i API-svar."
                time.sleep(1.0)
                continue
            content = (
                choices[0].get("message", {}).get("content")
                if isinstance(choices[0], dict)
                else ""
            )
            return str(content or "").strip()
        except Exception as e:
            last_err = str(e)
            time.sleep(1.0)
    raise RuntimeError(last_err)


class V2AIWorker(threading.Thread):
    def __init__(self, session_id: str, step1_only: bool = False, hint: str = ""):
        super().__init__(daemon=True)
        self.session_id = session_id
        self.step1_only = step1_only
        self.hint = hint.strip()

    def _session(self) -> Optional[V2Session]:
        with _sessions_lock:
            return _sessions.get(self.session_id)

    def _wait_if_paused(self, session: V2Session) -> bool:
        while session.ai_paused and not session.ai_stop_event.is_set():
            time.sleep(0.2)
        return not session.ai_stop_event.is_set()

    def _non_ovrigt_categories(self, session: V2Session) -> List[str]:
        return [c.name for c in session.categories if c.name.lower() != "övrigt"]

    def _examples_for_category(self, session: V2Session, category: str) -> List[str]:
        arts = [
            str(r.get("article_number", ""))
            for r in session.results
            if str(r.get("category", "")).strip().lower() == category.lower()
        ]
        uniq = []
        seen = set()
        for art in arts:
            if not art or art in seen:
                continue
            seen.add(art)
            uniq.append(art)
        return uniq[:MAX_EXAMPLES_PER_CAT]

    def _generate_knowledge_for_category(self, session: V2Session, category: str) -> str:
        examples = self._examples_for_category(session, category)
        session.cat_example_articles[category] = examples
        ai_url = str(session.ai_settings.get("api_url", "") or "").strip()
        model = str(session.ai_settings.get("model", "") or "").strip()
        api_key = str(session.ai_settings.get("api_key", "") or "").strip()
        if not ai_url or not model:
            if examples:
                return (
                    f"Kategori: {category}\n"
                    f"Exempelartiklar: {', '.join(examples)}\n"
                    "Använd kategoriens namn och exemplen som primär vägledning."
                )
            return f"Kategori: {category}\nIngen AI-analys tillgänglig ännu."

        sys_msg = (
            "Du hjälper till att skriva intern kategorikunskap för artikelklassificering. "
            "Svara kort, konkret och på svenska."
        )
        user_msg = (
            f"Syfte: {session.syfte}\n"
            f"Kategori: {category}\n"
            f"Exempelartiklar: {', '.join(examples) if examples else '(inga)'}\n\n"
            "Skriv en sammanfattning av vad som tillhör kategorin, vad som inte tillhör den, "
            "och viktiga tumregler."
        )
        return _call_chat_completion(
            api_url=ai_url,
            model=model,
            messages=[{"role": "system", "content": sys_msg}, {"role": "user", "content": user_msg}],
            api_key=api_key,
            timeout_sec=120,
            retries=3,
        )

    def _classify_article(self, session: V2Session, row: Dict[str, Any]) -> Tuple[str, str]:
        art = str(row.get("article_number", "")).strip()
        categories = session.category_names()
        meta = _current_data_manager().get_meta(art, str(row.get("bolag", ""))) or {}
        ai_url = str(session.ai_settings.get("api_url", "") or "").strip()
        model = str(session.ai_settings.get("model", "") or "").strip()
        api_key = str(session.ai_settings.get("api_key", "") or "").strip()

        if not ai_url or not model:
            hay = " ".join(
                [
                    str(meta.get("beskrivning", "")),
                    str(meta.get("kategori", "")),
                    str(meta.get("huvudkategori", "")),
                ]
            ).lower()
            for c in categories:
                if c.lower() == "övrigt":
                    continue
                if c.lower() in hay:
                    return c, "Heuristisk matchning på metadata."
            return "Övrigt", "Fallback till Övrigt (ingen AI-konfiguration)."

        knowledge_lines = []
        for c in categories:
            k = session.cat_knowledge.get(c, "")
            knowledge_lines.append(f"- {c}: {k[:1200]}")
        user_prompt = (
            "Klassificera en artikel till exakt EN kategori.\n"
            f"Tillåtna kategorier: {', '.join(categories)}\n\n"
            f"Artikelnummer: {art}\n"
            f"Bolag: {row.get('bolag', '')}\n"
            f"URL: {row.get('url', '')}\n"
            f"Beskrivning: {meta.get('beskrivning', '')}\n"
            f"Intern kategori: {meta.get('kategori', '')}\n"
            f"Huvudkategori: {meta.get('huvudkategori', '')}\n"
            f"Hint: {self.hint}\n\n"
            "Kategorikunskap:\n"
            + "\n".join(knowledge_lines)
            + "\n\nSvar ska vara JSON på formen: {\"category\":\"...\",\"reason\":\"...\"}"
        )
        raw = _call_chat_completion(
            api_url=ai_url,
            model=model,
            messages=[
                {"role": "system", "content": "Du returnerar endast giltig JSON."},
                {"role": "user", "content": user_prompt},
            ],
            api_key=api_key,
            timeout_sec=120,
            retries=3,
        )
        category = ""
        reason = ""
        try:
            parsed = json.loads(raw)
            category = str(parsed.get("category", "")).strip()
            reason = str(parsed.get("reason", "")).strip()
        except Exception:
            m = re.search(r'"category"\s*:\s*"([^"]+)"', raw)
            if m:
                category = m.group(1).strip()
            m2 = re.search(r'"reason"\s*:\s*"([^"]+)"', raw)
            if m2:
                reason = m2.group(1).strip()
        if category not in categories:
            category = "Övrigt" if "Övrigt" in categories else categories[0]
        if not reason:
            reason = "AI-klassificering."
        return category, reason

    def run(self):
        session = self._session()
        if not session:
            return
        with session.lock:
            if session.ai_running:
                return
            session.ai_running = True
            session.ai_error = ""
            session.ai_step = "step1"
            session.ai_stop_event.clear()
            session.ai_paused = False
            session.ai_pause_event.clear()
        _log(session, "=== Steg 1: Genererar kategorikunskap ===")

        try:
            cats = self._non_ovrigt_categories(session)
            if not cats:
                _log(session, "Inga kategorier att analysera i steg 1 (endast Övrigt finns).")
            for cat in cats:
                if session.ai_stop_event.is_set():
                    with session.lock:
                        session.ai_step = "stopped"
                    _log(session, "AI-jobb stoppat under steg 1.")
                    return
                if not self._wait_if_paused(session):
                    with session.lock:
                        session.ai_step = "stopped"
                    _log(session, "AI-jobb stoppat.")
                    return
                _log(session, f"Analyserar kategori: {cat}")
                knowledge = self._generate_knowledge_for_category(session, cat)
                with session.lock:
                    session.cat_knowledge[cat] = knowledge
                _log(session, f"Kategorikunskap klar: {cat}")

            with session.lock:
                if session.ai_stop_event.is_set():
                    session.ai_step = "stopped"
                    session.ai_running = False
                    return
                session.ai_step = "step1_done"
                session.ai_paused = True
            _log(session, "✓ Kategorikunskap klar — väntar på start av steg 2.")

            if self.step1_only:
                with session.lock:
                    session.ai_running = False
                return

            while session.ai_paused and not session.ai_stop_event.is_set():
                time.sleep(0.2)

            if session.ai_stop_event.is_set():
                with session.lock:
                    session.ai_step = "stopped"
                    session.ai_running = False
                _log(session, "AI-jobb stoppat före steg 2.")
                return

            with session.lock:
                session.ai_step = "step2"
            remaining = _unclassified_rows(session)
            with session.lock:
                session.ai_progress_done = 0
                session.ai_progress_total = len(remaining)
            _log(session, f"=== Steg 2: Klassificerar återstående artiklar ({len(remaining)}) ===")
            for row in remaining:
                if session.ai_stop_event.is_set():
                    with session.lock:
                        session.ai_step = "stopped"
                    _log(session, "AI-jobb stoppat under steg 2.")
                    return
                if not self._wait_if_paused(session):
                    with session.lock:
                        session.ai_step = "stopped"
                    _log(session, "AI-jobb stoppat.")
                    return
                art = str(row.get("article_number", "")).strip()
                try:
                    cat, reason = self._classify_article(session, row)
                except Exception as e:
                    cat = "Övrigt" if "Övrigt" in session.category_names() else session.category_names()[0]
                    reason = f"Fallback: {e}"
                with session.lock:
                    _set_result(session, row, cat, reason=reason, source="ai")
                    session.ai_progress_done += 1
                _log(session, f"[{session.ai_progress_done}/{session.ai_progress_total}] {art} -> {cat}")

            with session.lock:
                session.ai_step = "done"
                session.ai_running = False
                session.ai_paused = False
                _advance_to_next_unclassified(session)
                if len(_result_map(session)) >= len(session.csv_data) and session.csv_data:
                    session.stage = "done"
            _log(session, "✓ AI-jobb klart.")
        except Exception as e:
            with session.lock:
                session.ai_step = "error"
                session.ai_error = str(e)
                session.ai_running = False
                session.ai_paused = False
            _log(session, f"FEL i AI-jobb: {e}")
        finally:
            with _active_workers_lock:
                _active_workers.pop(self.session_id, None)


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------


class V2SessionCreateBody(BaseModel):
    test_name: str = ""
    syfte: str = DEFAULT_SYFTE
    categories: List[Any] = []
    image_dir: str = str(DEFAULT_IMAGE_DIR)
    output_dir: str = str(DEFAULT_OUTPUT_DIR)
    shuffle: bool = False
    source_mode: str = "auto"  # auto|images|item_attribute
    ai_settings: Dict[str, Any] = {}


class V2ClassifyBody(BaseModel):
    category: str = ""
    reason: str = ""


class V2ReclassifyBody(BaseModel):
    article_numbers: List[str] = []
    target_category: str = ""
    reason: str = ""
    use_ai: bool = False
    hint: str = ""


class V2AddCategoryBody(BaseModel):
    name: str = ""
    description: str = ""


class V2KnowledgeUpdateBody(BaseModel):
    category: str = ""
    new_name: str = ""
    knowledge: str = ""
    example_articles: List[str] = []


class V2ActionBody(BaseModel):
    action: str = ""
    payload: Dict[str, Any] = {}


# ---------------------------------------------------------------------------
# Config + data files
# ---------------------------------------------------------------------------


@router.get("/config")
def api_v2_config():
    return {
        "ok": True,
        "version": "v2",
        "default_test_name": "test1",
        "default_syfte": DEFAULT_SYFTE,
        "default_image_dir": str(DEFAULT_IMAGE_DIR),
        "default_output_dir": str(DEFAULT_OUTPUT_DIR),
        "default_categories": [
            {"name": "Hund", "description": ""},
            {"name": "Katt", "description": ""},
            {"name": "Övrigt", "description": "Fallbackkategori"},
        ],
        "data_files": _data_files_payload(),
        "ai_defaults": {
            "api_url": DEFAULT_AI_URL,
            "model": "",
            "compress_images": True,
            "batch_size": EXT_BATCH_SIZE,
            "parallel_workers": AI_PARALLEL_WORKERS,
        },
    }


@router.get("/data-files")
def api_v2_data_files():
    item_rows = 0
    with _data_files_lock:
        path = _data_files.get("item_attribute", "")
    if path and os.path.exists(path):
        item_rows = len(_extract_rows_from_item_attribute(path))
    return {"ok": True, "files": _data_files_payload(), "item_attribute_rows": item_rows}


@router.post("/data-files/upload")
async def api_v2_data_upload(file_key: str = Form(...), file: UploadFile = File(...)):
    if file_key not in CLASSIFIER_DATA_FILE_KEYS:
        raise HTTPException(status_code=400, detail=f"Ogiltig file_key: {file_key}")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Tom fil.")
    base = _resolve_data_dir()
    safe = _safe_name(file.filename or f"{file_key}.csv")
    dest = base / f"{file_key}_{safe}"
    with open(dest, "wb") as fh:
        fh.write(content)
    old = ""
    with _data_files_lock:
        old = _data_files.get(file_key, "")
        _data_files[file_key] = str(dest)
    if old and old != str(dest):
        try:
            if os.path.exists(old):
                os.remove(old)
        except Exception:
            pass
    item_rows = 0
    if file_key == "item_attribute":
        item_rows = len(_extract_rows_from_item_attribute(str(dest)))
    return {
        "ok": True,
        "file_key": file_key,
        "filename": os.path.basename(dest),
        "bytes": len(content),
        "item_attribute_rows": item_rows,
        "files": _data_files_payload(),
    }


@router.delete("/data-files/{file_key}")
def api_v2_data_delete(file_key: str):
    if file_key not in CLASSIFIER_DATA_FILE_KEYS:
        raise HTTPException(status_code=400, detail=f"Ogiltig file_key: {file_key}")
    old = ""
    with _data_files_lock:
        old = _data_files.pop(file_key, "")
    if old and os.path.exists(old):
        try:
            os.remove(old)
        except Exception:
            pass
    return {"ok": True, "files": _data_files_payload()}


# ---------------------------------------------------------------------------
# Session lifecycle
# ---------------------------------------------------------------------------


@router.post("/session/start")
def api_v2_session_start(body: V2SessionCreateBody):
    test_name = str(body.test_name or "").strip()
    if not test_name:
        raise HTTPException(status_code=400, detail="Testnamn krävs.")
    categories = _normalize_categories(body.categories)
    if not categories:
        raise HTTPException(status_code=400, detail="Minst en kategori krävs.")

    output_dir = Path((body.output_dir or str(DEFAULT_OUTPUT_DIR)).strip()).expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)

    source_mode = str(body.source_mode or "auto").strip().lower()
    image_dir = Path((body.image_dir or str(DEFAULT_IMAGE_DIR)).strip()).expanduser()
    rows: List[Dict[str, Any]] = []
    image_source = "url"

    if source_mode in {"auto", "images"}:
        rows = _rows_from_image_dir(image_dir)
        if rows:
            image_source = "image_dir"
    if not rows and source_mode in {"auto", "item_attribute"}:
        item_attr_path = ""
        with _data_files_lock:
            item_attr_path = _data_files.get("item_attribute", "")
        if item_attr_path and os.path.exists(item_attr_path):
            rows = _extract_rows_from_item_attribute(item_attr_path)
            image_source = "url"
    if not rows:
        raise HTTPException(
            status_code=400,
            detail=(
                "Inga artiklar hittades. Kontrollera bildmapp eller ladda upp item_attribute "
                "med IMG-URL."
            ),
        )
    if body.shuffle:
        random.shuffle(rows)

    sid = uuid.uuid4().hex[:12]
    session = V2Session(
        session_id=sid,
        test_name=test_name,
        syfte=str(body.syfte or DEFAULT_SYFTE),
        categories=categories,
        csv_data=rows,
        output_dir=str(output_dir.resolve()),
        image_dir=str(image_dir),
        image_source=image_source,
        ai_settings=dict(body.ai_settings or {}),
    )
    _log(session, f"Session startad ({len(rows)} artiklar, källa: {image_source}).")
    with _sessions_lock:
        _sessions[sid] = session
    return _session_state(session)


@router.get("/session/{session_id}")
def api_v2_session_state(session_id: str):
    return _session_state(_find_session_or_404(session_id))


@router.delete("/session/{session_id}")
def api_v2_session_delete(session_id: str):
    session = _find_session_or_404(session_id)
    with _sessions_lock:
        _sessions.pop(session_id, None)
    try:
        shutil.rmtree(session.temp_dir, ignore_errors=True)
    except Exception:
        pass
    return {"ok": True}


# ---------------------------------------------------------------------------
# Core actions
# ---------------------------------------------------------------------------


@router.post("/session/{session_id}/classify")
def api_v2_classify(session_id: str, body: V2ClassifyBody):
    session = _find_session_or_404(session_id)
    category = str(body.category or "").strip()
    if not category:
        raise HTTPException(status_code=400, detail="Kategori krävs.")
    with session.lock:
        _advance_to_next_unclassified(session)
        row = _current_row(session)
        if not row:
            session.stage = "done"
            return _session_state(session)
        _set_result(session, row, category, reason=body.reason or "", source="manual")
        art = str(row.get("article_number", "")).strip()
        _log(session, f"Klassificerad: {art} -> {category}")
        session.current_index += 1
        _advance_to_next_unclassified(session)
        if len(_result_map(session)) >= len(session.csv_data) and session.csv_data:
            session.stage = "done"
    return _session_state(session)


@router.post("/session/{session_id}/skip")
def api_v2_skip(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        _advance_to_next_unclassified(session)
        row = _current_row(session)
        if not row:
            session.stage = "done"
            return _session_state(session)
        art = str(row.get("article_number", "")).strip()
        if art and art not in session.skipped:
            session.skipped.append(art)
        session.current_index += 1
        _advance_to_next_unclassified(session)
        _log(session, f"Hoppade över: {art}")
    return _session_state(session)


@router.post("/session/{session_id}/back")
def api_v2_back(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        if session.current_index > 0:
            session.current_index -= 1
        if session.stage == "done":
            session.stage = "classify"
        row = _current_row(session)
        _log(session, f"Gick tillbaka till index {session.current_index}.")
        if not row and session.csv_data:
            session.current_index = max(0, len(session.csv_data) - 1)
    return _session_state(session)


@router.post("/session/{session_id}/finish")
def api_v2_finish(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        session.stage = "done"
        _log(session, "Session avslutad i förtid.")
    return _session_state(session)


@router.post("/session/{session_id}/add-category")
def api_v2_add_category(session_id: str, body: V2AddCategoryBody):
    session = _find_session_or_404(session_id)
    name = str(body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Kategorinamn krävs.")
    with session.lock:
        existing = {c.name.lower() for c in session.categories}
        if name.lower() in existing:
            raise HTTPException(status_code=400, detail=f"Kategorin finns redan: {name}")
        session.categories.append(V2Category(name=name, description=str(body.description or "").strip()))
        _log(session, f"Ny kategori skapad: {name}")
    return _session_state(session)


@router.post("/session/{session_id}/reclassify")
def api_v2_reclassify(session_id: str, body: V2ReclassifyBody):
    session = _find_session_or_404(session_id)
    numbers = [str(x).strip() for x in (body.article_numbers or []) if str(x).strip()]
    if not numbers:
        raise HTTPException(status_code=400, detail="article_numbers saknas.")
    target = str(body.target_category or "").strip()
    if not target:
        raise HTTPException(status_code=400, detail="target_category krävs.")

    with session.lock:
        if target not in session.category_names():
            raise HTTPException(status_code=400, detail=f"Okänd kategori: {target}")
        res_map = _result_map(session)
        changed = 0
        for art in numbers:
            res = res_map.get(art)
            if not res:
                continue
            res["category"] = target
            if body.reason:
                res["reason"] = body.reason
            changed += 1
        session.results = list(res_map.values())
        cat_map = {str(x.get("article_number", "")).strip(): x for x in session.categorized}
        for art in numbers:
            if art in cat_map:
                cat_map[art]["category"] = target
        session.categorized = list(cat_map.values())
        _log(
            session,
            f"Omklassificerade {changed} artikel{'er' if changed != 1 else ''} till {target}."
            + (f" Orsak: {body.reason}" if body.reason else ""),
        )
    return _session_state(session)


@router.post("/session/{session_id}/knowledge/update")
def api_v2_knowledge_update(session_id: str, body: V2KnowledgeUpdateBody):
    session = _find_session_or_404(session_id)
    category = str(body.category or "").strip()
    if not category:
        raise HTTPException(status_code=400, detail="category krävs.")
    with session.lock:
        names = session.category_names()
        if category not in names:
            raise HTTPException(status_code=400, detail=f"Okänd kategori: {category}")
        new_name = str(body.new_name or "").strip()
        if new_name and new_name != category:
            if new_name in names:
                raise HTTPException(status_code=400, detail=f"Kategorin finns redan: {new_name}")
            for c in session.categories:
                if c.name == category:
                    c.name = new_name
                    break
            for r in session.results:
                if str(r.get("category", "")) == category:
                    r["category"] = new_name
            for c in session.categorized:
                if str(c.get("category", "")) == category:
                    c["category"] = new_name
            if category in session.cat_knowledge:
                session.cat_knowledge[new_name] = session.cat_knowledge.pop(category)
            if category in session.cat_example_articles:
                session.cat_example_articles[new_name] = session.cat_example_articles.pop(category)
            category = new_name
            _log(session, f"Kategori namnbytt: {body.category} -> {new_name}")

        if body.knowledge:
            session.cat_knowledge[category] = body.knowledge
        if body.example_articles is not None:
            clean = [str(x).strip() for x in body.example_articles if str(x).strip()]
            session.cat_example_articles[category] = clean[:MAX_EXAMPLES_PER_CAT]
        _log(session, f"Kategorikunskap uppdaterad: {category}")
    return _session_state(session)


# ---------------------------------------------------------------------------
# AI controls
# ---------------------------------------------------------------------------


def _start_worker(session: V2Session, worker: V2AIWorker):
    with _active_workers_lock:
        existing = _active_workers.get(session.session_id)
        if existing and existing.is_alive():
            raise HTTPException(status_code=409, detail="AI-jobb kör redan.")
        _active_workers[session.session_id] = worker
    worker.start()


@router.post("/session/{session_id}/ai/start")
def api_v2_ai_start(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        session.ai_stop_event.clear()
        session.ai_paused = False
        session.ai_step = "step1"
    _log(session, "AI-jobb initierat.")
    _start_worker(session, V2AIWorker(session_id=session_id, step1_only=False))
    return _session_state(session)


@router.post("/session/{session_id}/ai/analyze-all")
def api_v2_ai_analyze_all(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        session.ai_stop_event.clear()
        session.ai_paused = False
        session.ai_step = "step1"
    _log(session, "Analysera om alla kategorier startat.")
    _start_worker(session, V2AIWorker(session_id=session_id, step1_only=True))
    return _session_state(session)


@router.post("/session/{session_id}/ai/start-step2")
def api_v2_ai_start_step2(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        session.ai_paused = False
        if session.ai_step == "step1_done":
            _log(session, "Steg 2 startad.")
        elif not session.ai_running:
            raise HTTPException(status_code=400, detail="Inget AI-jobb väntar på steg 2.")
    return _session_state(session)


@router.post("/session/{session_id}/ai/pause")
def api_v2_ai_pause(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        if not session.ai_running:
            raise HTTPException(status_code=400, detail="AI-jobb kör inte.")
        session.ai_paused = True
        _log(session, "AI-jobb pausat.")
    return _session_state(session)


@router.post("/session/{session_id}/ai/resume")
def api_v2_ai_resume(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        if not session.ai_running and session.ai_step != "step1_done":
            raise HTTPException(status_code=400, detail="AI-jobb kör inte.")
        session.ai_paused = False
        _log(session, "AI-jobb återupptaget.")
    return _session_state(session)


@router.post("/session/{session_id}/ai/stop")
def api_v2_ai_stop(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        session.ai_stop_event.set()
        session.ai_paused = False
        _log(session, "Stoppsignal skickad till AI-jobb.")
    return _session_state(session)


# ---------------------------------------------------------------------------
# Generic action endpoint
# ---------------------------------------------------------------------------


@router.post("/session/{session_id}/action")
def api_v2_action(session_id: str, body: V2ActionBody):
    action = str(body.action or "").strip().lower()
    payload = body.payload or {}
    if action == "classify":
        return api_v2_classify(session_id, V2ClassifyBody(**payload))
    if action == "skip":
        return api_v2_skip(session_id)
    if action == "back":
        return api_v2_back(session_id)
    if action == "finish":
        return api_v2_finish(session_id)
    if action == "add_category":
        return api_v2_add_category(session_id, V2AddCategoryBody(**payload))
    if action == "reclassify":
        return api_v2_reclassify(session_id, V2ReclassifyBody(**payload))
    if action == "knowledge_update":
        return api_v2_knowledge_update(session_id, V2KnowledgeUpdateBody(**payload))
    if action == "ai_start":
        return api_v2_ai_start(session_id)
    if action == "ai_pause":
        return api_v2_ai_pause(session_id)
    if action == "ai_resume":
        return api_v2_ai_resume(session_id)
    if action == "ai_stop":
        return api_v2_ai_stop(session_id)
    if action == "ai_start_step2":
        return api_v2_ai_start_step2(session_id)
    if action == "ai_analyze_all":
        return api_v2_ai_analyze_all(session_id)
    raise HTTPException(status_code=400, detail=f"Okänd action: {action}")


# ---------------------------------------------------------------------------
# Image + event stream
# ---------------------------------------------------------------------------


@router.get("/session/{session_id}/image")
def api_v2_image(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        _advance_to_next_unclassified(session)
        row = _current_row(session)
    if not row:
        raise HTTPException(status_code=404, detail="Ingen aktiv artikel.")

    img_path = str(row.get("img_path", "") or "").strip()
    if img_path and Path(img_path).exists():
        p = Path(img_path)
        media_type = mimetypes.guess_type(str(p))[0] or "application/octet-stream"
        return FileResponse(str(p), media_type=media_type, filename=p.name)

    url = str(row.get("url", "") or "").strip()
    if not url:
        raise HTTPException(status_code=404, detail="Bilden saknar både img_path och url.")
    try:
        content, content_type = _download_url_bytes(url, timeout_sec=45)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Kunde inte hämta bild från URL: {e}") from e
    return Response(content=content, media_type=content_type)


@router.get("/session/{session_id}/events")
async def api_v2_events(session_id: str):
    session = _find_session_or_404(session_id)

    async def event_gen():
        while True:
            try:
                msg = await asyncio.to_thread(session.event_queue.get, True, 25)
                payload = json.dumps({"type": "log", "message": msg}, ensure_ascii=False)
                yield f"data: {payload}\n\n"
            except queue.Empty:
                yield ": keepalive\n\n"
            except asyncio.CancelledError:
                break
            except Exception:
                break

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


# ---------------------------------------------------------------------------
# Import/export parity (ZIP + Excel)
# ---------------------------------------------------------------------------


def _build_zip_payload_files(session: V2Session) -> Tuple[Dict[str, Any], List[Tuple[str, str]]]:
    img_srcs: Dict[str, str] = {}

    def _register(src: str):
        if not src or src in img_srcs:
            return
        p = Path(src)
        if not p.exists():
            return
        name = p.name
        taken = set(img_srcs.values())
        if f"images/{name}" in taken:
            i = 1
            while f"images/{p.stem}_{i}{p.suffix}" in taken:
                i += 1
            name = f"{p.stem}_{i}{p.suffix}"
        img_srcs[src] = f"images/{name}"

    for item in session.categorized:
        _register(str(item.get("image_path", "") or ""))
    for row in session.csv_data:
        _register(str(row.get("img_path", "") or ""))

    def _rel(src: str) -> str:
        return img_srcs.get(src, src)

    sess = {
        "test_name": session.test_name,
        "syfte": session.syfte,
        "categories": [c.to_dict() for c in session.categories],
        "cat_knowledge": session.cat_knowledge,
        "cat_example_articles": session.cat_example_articles,
    }
    csv_export = [{**r, "img_path": _rel(str(r.get("img_path", "") or ""))} for r in session.csv_data]
    cat_export = [{**c, "image_path": _rel(str(c.get("image_path", "") or ""))} for c in session.categorized]
    files = [(orig, arc) for orig, arc in img_srcs.items()]
    return {"session": sess, "csv_data": csv_export, "categorized": cat_export, "results": session.results}, files


@router.get("/session/{session_id}/export/zip")
def api_v2_export_zip(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        payload, files = _build_zip_payload_files(session)
    fd, tmp_path = tempfile.mkstemp(prefix="classifier_v2_", suffix=".zip")
    os.close(fd)
    try:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("session.json", json.dumps(payload["session"], ensure_ascii=False, indent=2))
            zf.writestr("csv_data.json", json.dumps(payload["csv_data"], ensure_ascii=False, indent=2))
            zf.writestr("categorized.json", json.dumps(payload["categorized"], ensure_ascii=False, indent=2))
            if payload["results"]:
                zf.writestr("results.json", json.dumps(payload["results"], ensure_ascii=False, indent=2))
            for orig, arc in files:
                zf.write(orig, arc)
    except Exception:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        raise
    name = f"{_safe_folder_fragment(session.test_name)}_session_{_now_token()}.zip"
    return FileResponse(tmp_path, media_type="application/zip", filename=name)


@router.post("/session/import/zip")
async def api_v2_import_zip(file: UploadFile = File(...)):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Tom ZIP-fil.")
    extract_dir = Path(tempfile.mkdtemp(prefix="classifier_v2_import_"))
    zip_path = extract_dir / "session.zip"
    with open(zip_path, "wb") as fh:
        fh.write(content)

    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(extract_dir)

        def _load(name: str, default):
            p = extract_dir / name
            if p.exists():
                with open(p, encoding="utf-8") as fh:
                    return json.load(fh)
            return default

        session_json = _load("session.json", {})
        csv_data = _load("csv_data.json", [])
        categorized = _load("categorized.json", [])
        results = _load("results.json", [])

        def _fix_img(p: str) -> str:
            if not p:
                return p
            full = extract_dir / p
            return str(full) if full.exists() else p

        for row in csv_data:
            row["img_path"] = _fix_img(str(row.get("img_path", "") or ""))
        for row in categorized:
            row["image_path"] = _fix_img(str(row.get("image_path", "") or ""))

        categories_raw = session_json.get("categories", [])
        categories = _normalize_categories(categories_raw)
        sid = uuid.uuid4().hex[:12]
        sess = V2Session(
            session_id=sid,
            test_name=str(session_json.get("test_name", "Import")),
            syfte=str(session_json.get("syfte", DEFAULT_SYFTE)),
            categories=categories,
            csv_data=csv_data,
            output_dir=str(DEFAULT_OUTPUT_DIR),
            image_dir="",
            image_source="mixed",
            results=results if isinstance(results, list) else [],
            categorized=categorized if isinstance(categorized, list) else [],
            cat_knowledge=dict(session_json.get("cat_knowledge", {}) or {}),
            cat_example_articles=dict(session_json.get("cat_example_articles", {}) or {}),
            temp_dir=str(extract_dir),
        )
        _advance_to_next_unclassified(sess)
        if len(_result_map(sess)) >= len(sess.csv_data) and sess.csv_data:
            sess.stage = "done"
        _log(sess, "ZIP-session importerad.")
        with _sessions_lock:
            _sessions[sid] = sess
        return _session_state(sess)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Kunde inte läsa ZIP: {e}") from e


def _excel_workbook_from_session(session: V2Session) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultat"
    headers = [
        "Artikelnummer",
        "Resultat kategori",
        "Huvudkategori",
        "Beskrivning",
        "Längd (mm)",
        "Bredd (mm)",
        "Höjd (mm)",
        "Volym",
        "Vikt brutto (kg)",
        "Vikt netto (kg)",
        "Robot (Y/N)",
        "StoreQuantity",
        "Bild (URL)",
        "Bolag",
        "Reason",
    ]
    ws.append(headers)
    dm = _current_data_manager()
    for row in session.results:
        art = str(row.get("article_number", ""))
        meta = dm.get_meta(art, str(row.get("bolag", ""))) or {}
        ws.append(
            [
                art,
                row.get("category", ""),
                meta.get("huvudkategori", ""),
                meta.get("beskrivning", ""),
                meta.get("langd", ""),
                meta.get("bredd", ""),
                meta.get("hojd", ""),
                meta.get("volym", ""),
                meta.get("vikt_brutto", ""),
                meta.get("vikt_netto", ""),
                meta.get("robot", ""),
                meta.get("store_quantity", ""),
                row.get("url", ""),
                row.get("bolag", ""),
                row.get("reason", ""),
            ]
        )

    col_widths = [20, 25, 25, 40, 12, 12, 12, 12, 18, 18, 12, 15, 60, 12, 45]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    classified = set(_result_map(session).keys())
    unclassified = [
        r for r in session.csv_data if str(r.get("article_number", "")).strip() not in classified
    ]
    if unclassified:
        ws_u = wb.create_sheet("Oklassificerade")
        ws_u.append(["Artikelnummer", "Bild (URL)", "Bolag"])
        for row in unclassified:
            ws_u.append([row.get("article_number", ""), row.get("url", ""), row.get("bolag", "")])
        ws_u.column_dimensions["A"].width = 20
        ws_u.column_dimensions["B"].width = 60
        ws_u.column_dimensions["C"].width = 15

    ws_s = wb.create_sheet("Session")
    ws_s.append(["Nyckel", "Värde"])
    ws_s.append(["test_name", session.test_name])
    ws_s.append(["syfte", session.syfte])
    ws_s.append(["categories_json", json.dumps([c.to_dict() for c in session.categories], ensure_ascii=False)])
    if session.cat_knowledge:
        ws_s.append(["cat_knowledge_json", json.dumps(session.cat_knowledge, ensure_ascii=False)])
    if session.cat_example_articles:
        ws_s.append(
            [
                "cat_example_articles_json",
                json.dumps(session.cat_example_articles, ensure_ascii=False),
            ]
        )
    ws_s.column_dimensions["A"].width = 30
    ws_s.column_dimensions["B"].width = 120

    if session.cat_knowledge:
        ws_k = wb.create_sheet("Kategorianalys")
        ws_k.append(["Kategori", "AI-analys", "Exempelartiklar"])
        for cat_name, knowledge in session.cat_knowledge.items():
            example_arts = session.cat_example_articles.get(cat_name, [])
            ws_k.append([cat_name, knowledge, ", ".join(example_arts)])
        ws_k.column_dimensions["A"].width = 25
        ws_k.column_dimensions["B"].width = 85
        ws_k.column_dimensions["C"].width = 45

    fd, tmp_path = tempfile.mkstemp(prefix="classifier_v2_", suffix=".xlsx")
    os.close(fd)
    wb.save(tmp_path)
    return tmp_path


@router.get("/session/{session_id}/export/excel")
def api_v2_export_excel(session_id: str):
    session = _find_session_or_404(session_id)
    with session.lock:
        path = _excel_workbook_from_session(session)
    name = f"{_safe_folder_fragment(session.test_name)}_resultat_{_now_token()}.xlsx"
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=name,
    )


@router.post("/session/import/excel")
async def api_v2_import_excel(file: UploadFile = File(...)):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Tom Excel-fil.")
    fd, tmp_path = tempfile.mkstemp(prefix="classifier_v2_import_", suffix=".xlsx")
    os.close(fd)
    with open(tmp_path, "wb") as fh:
        fh.write(content)

    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        session_data: Dict[str, str] = {}
        if "Session" in wb.sheetnames:
            for row in wb["Session"].iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2 and row[0] and row[1] is not None:
                    session_data[str(row[0])] = str(row[1])

        test_name = session_data.get("test_name", Path(file.filename or tmp_path).stem)
        syfte = session_data.get("syfte", DEFAULT_SYFTE)
        try:
            categories_json = json.loads(session_data.get("categories_json", "[]"))
        except Exception:
            categories_json = []
        categories = _normalize_categories(categories_json)
        try:
            cat_knowledge = json.loads(session_data.get("cat_knowledge_json", "{}"))
        except Exception:
            cat_knowledge = {}
        try:
            cat_example_articles = json.loads(session_data.get("cat_example_articles_json", "{}"))
        except Exception:
            cat_example_articles = {}

        results: List[Dict[str, Any]] = []
        csv_data: List[Dict[str, Any]] = []
        categorized: List[Dict[str, Any]] = []

        if "Resultat" in wb.sheetnames:
            ws = wb["Resultat"]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(x).strip() if x is not None else "" for x in header_row]
            hmap = {v: i for i, v in enumerate(headers) if v}

            def _cell(row_vals, key, default=""):
                idx = hmap.get(key)
                if idx is None or idx >= len(row_vals):
                    return default
                val = row_vals[idx]
                return str(val).strip() if val is not None else default

            for row_vals in ws.iter_rows(min_row=2, values_only=True):
                if not any(row_vals):
                    continue
                art = _cell(row_vals, "Artikelnummer")
                cat = _cell(row_vals, "Resultat kategori")
                url = _cell(row_vals, "Bild (URL)")
                bolag = _cell(row_vals, "Bolag")
                reason = _cell(row_vals, "Reason")
                if not art:
                    continue
                results.append(
                    {
                        "article_number": art,
                        "category": cat or "Övrigt",
                        "url": url,
                        "bolag": bolag,
                        "reason": reason,
                        "image_path": "",
                        "source": "import",
                    }
                )
                categorized.append({"article_number": art, "category": cat or "Övrigt", "image_path": ""})
                csv_data.append({"article_number": art, "url": url, "bolag": bolag, "img_path": ""})

        if "Oklassificerade" in wb.sheetnames:
            ws_u = wb["Oklassificerade"]
            u_header = next(ws_u.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(x).strip() if x is not None else "" for x in u_header]
            hmap = {v: i for i, v in enumerate(headers) if v}

            def _ucell(row_vals, key, default=""):
                idx = hmap.get(key)
                if idx is None or idx >= len(row_vals):
                    return default
                val = row_vals[idx]
                return str(val).strip() if val is not None else default

            for row_vals in ws_u.iter_rows(min_row=2, values_only=True):
                if not any(row_vals):
                    continue
                art = _ucell(row_vals, "Artikelnummer")
                url = _ucell(row_vals, "Bild (URL)")
                bolag = _ucell(row_vals, "Bolag")
                if not art:
                    continue
                csv_data.append({"article_number": art, "url": url, "bolag": bolag, "img_path": ""})

        wb.close()
        sid = uuid.uuid4().hex[:12]
        sess = V2Session(
            session_id=sid,
            test_name=test_name,
            syfte=syfte,
            categories=categories,
            csv_data=csv_data,
            output_dir=str(DEFAULT_OUTPUT_DIR),
            image_source="url",
            results=results,
            categorized=categorized,
            cat_knowledge=dict(cat_knowledge or {}),
            cat_example_articles=dict(cat_example_articles or {}),
        )
        _advance_to_next_unclassified(sess)
        if len(_result_map(sess)) >= len(sess.csv_data) and sess.csv_data:
            sess.stage = "done"
        _log(sess, "Excel-session importerad.")
        with _sessions_lock:
            _sessions[sid] = sess
        return _session_state(sess)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Kunde inte läsa Excel: {e}") from e


# ---------------------------------------------------------------------------
# Diagnostics
# ---------------------------------------------------------------------------


@router.get("/session/{session_id}/logs")
def api_v2_logs(session_id: str, tail: int = 200):
    session = _find_session_or_404(session_id)
    tail = max(1, min(2000, int(tail)))
    with session.lock:
        return {"ok": True, "logs": session.log_lines[-tail:]}


@router.get("/sessions")
def api_v2_list_sessions():
    out = []
    with _sessions_lock:
        for sid, session in _sessions.items():
            out.append(
                {
                    "session_id": sid,
                    "test_name": session.test_name,
                    "stage": session.stage,
                    "total": len(session.csv_data),
                    "classified": len(session.results),
                    "ai_step": session.ai_step,
                    "ai_running": session.ai_running,
                }
            )
    return {"ok": True, "sessions": out}
